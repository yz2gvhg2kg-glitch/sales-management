import uuid
import io
import re
from datetime import date, datetime
from typing import Optional, List

import openpyxl
from fastapi import UploadFile, HTTPException

from app.core.config import settings


def generate_order_no(prefix: str = "ORD") -> str:
    """Generate unique order number: ORD20260804A1B2C3D4"""
    short_uuid = uuid.uuid4().hex[:8].upper()
    return f"{prefix}{date.today().strftime('%Y%m%d')}{short_uuid}"


def generate_customer_no() -> str:
    """Generate unique customer number."""
    return f"CUS{date.today().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"


def sanitize_phone(phone: Optional[str]) -> Optional[str]:
    """Clean and validate phone number."""
    if not phone:
        return None
    cleaned = re.sub(r'[^\d]', '', phone)
    if len(cleaned) < 7 or len(cleaned) > 15:
        return None
    return cleaned


def validate_file_upload(file: UploadFile) -> None:
    """Validate uploaded file size and extension."""
    ext = file.filename.rsplit('.', 1)[-1].lower() if file.filename else ''
    allowed = [e.strip('.') for e in settings.ALLOWED_UPLOAD_EXTENSIONS.split(',')]
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"文件类型不支持，仅允许: {settings.ALLOWED_UPLOAD_EXTENSIONS}")


def parse_excel(file_content: bytes) -> tuple[list[str], list[dict]]:
    """Parse Excel file content into headers and rows."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    ws = wb.active
    headers = [str(cell.value or '') for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = value
        # Skip completely empty rows
        if any(v for v in row_data.values() if v is not None):
            rows.append(row_data)
    return headers, rows


def safe_int(value, default=0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def safe_float(value, default=0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_str(value, default="") -> str:
    if value is None:
        return default
    return str(value).strip()


def format_date(d: Optional[date]) -> Optional[str]:
    return d.isoformat() if d else None


def format_datetime(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() if dt else None

from typing import Optional
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
import uuid
import io

from app.core.database import get_db
from app.core.security import get_current_user, get_current_admin
from app.models.order import Order, OrderStatus, AfterSales
from app.models.user import User

router = APIRouter()


class OrderCreate(BaseModel):
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    product_name: str
    product_id: Optional[int] = None
    quantity: int = 1
    unit_price: float
    total_amount: float
    cost_amount: float = 0
    source: Optional[str] = None
    order_date: date
    remark: Optional[str] = None


class AfterSalesCreate(BaseModel):
    order_id: int
    type: str  # return/exchange/rejection
    reason: Optional[str] = None
    refund_amount: float = 0


@router.get("")
async def list_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    salesperson_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Order)
    count_query = select(func.count(Order.id))

    if current_user.role == "employee":
        query = query.where(Order.salesperson_id == current_user.id)
        count_query = count_query.where(Order.salesperson_id == current_user.id)
    elif salesperson_id:
        query = query.where(Order.salesperson_id == salesperson_id)
        count_query = count_query.where(Order.salesperson_id == salesperson_id)

    if status:
        query = query.where(Order.status == status)
        count_query = count_query.where(Order.status == status)
    if keyword:
        query = query.where(
            (Order.customer_name.contains(keyword))
            | (Order.order_no.contains(keyword))
            | (Order.product_name.contains(keyword))
        )
        count_query = count_query.where(
            (Order.customer_name.contains(keyword))
            | (Order.order_no.contains(keyword))
            | (Order.product_name.contains(keyword))
        )
    if start_date:
        query = query.where(Order.order_date >= start_date)
        count_query = count_query.where(Order.order_date >= start_date)
    if end_date:
        query = query.where(Order.order_date <= end_date)
        count_query = count_query.where(Order.order_date <= end_date)

    total = (await db.execute(count_query)).scalar()
    result = await db.execute(
        query.order_by(Order.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )
    orders = result.scalars().all()

    return {
        "total": total,
        "items": [
            {
                "id": o.id,
                "order_no": o.order_no,
                "customer_name": o.customer_name,
                "customer_phone": o.customer_phone,
                "product_name": o.product_name,
                "quantity": o.quantity,
                "unit_price": o.unit_price,
                "total_amount": o.total_amount,
                "cost_amount": o.cost_amount,
                "salesperson_id": o.salesperson_id,
                "source": o.source,
                "status": o.status.value,
                "order_date": o.order_date.isoformat() if o.order_date else None,
                "remark": o.remark,
                "created_at": o.created_at.isoformat() if o.created_at else None,
            }
            for o in orders
        ],
    }


@router.post("")
async def create_order(
    data: OrderCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    order_no = f"ORD{date.today().strftime('%Y%m%d')}{uuid.uuid4().hex[:8].upper()}"
    order = Order(
        order_no=order_no,
        customer_name=data.customer_name,
        customer_phone=data.customer_phone,
        product_id=data.product_id,
        product_name=data.product_name,
        quantity=data.quantity,
        unit_price=data.unit_price,
        total_amount=data.total_amount,
        cost_amount=data.cost_amount,
        salesperson_id=current_user.id,
        source=data.source,
        order_date=data.order_date,
        remark=data.remark,
    )
    db.add(order)
    await db.commit()
    await db.refresh(order)
    return {"id": order.id, "order_no": order_no, "message": "创建成功"}


@router.post("/import")
async def import_orders(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    """批量导入订单"""
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        order_no = f"ORD{date.today().strftime('%Y%m%d')}{uuid.uuid4().hex[:8].upper()}"
        order = Order(
            order_no=order_no,
            customer_name=str(row_data.get("客户姓名") or row_data.get("customer_name") or ""),
            customer_phone=str(row_data.get("电话") or row_data.get("phone") or ""),
            product_name=str(row_data.get("产品") or row_data.get("product_name") or ""),
            quantity=int(row_data.get("数量") or row_data.get("quantity") or 1),
            unit_price=float(row_data.get("单价") or row_data.get("unit_price") or 0),
            total_amount=float(row_data.get("总金额") or row_data.get("total_amount") or 0),
            cost_amount=float(row_data.get("成本") or row_data.get("cost") or 0),
            salesperson_id=int(row_data.get("业务员ID") or row_data.get("salesperson_id") or 0),
            source=row_data.get("来源") or row_data.get("source"),
            order_date=date.today(),
        )
        orders.append(order)

    db.add_all(orders)
    await db.commit()
    return {"message": f"成功导入 {len(orders)} 条订单"}


@router.post("/after-sales")
async def create_after_sales(
    data: AfterSalesCreate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    """创建售后记录（退货/换货/拒收）"""
    result = await db.execute(select(Order).where(Order.id == data.order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    after_sales = AfterSales(
        order_id=data.order_id,
        type=data.type,
        reason=data.reason,
        refund_amount=data.refund_amount,
    )
    db.add(after_sales)

    # 更新订单状态
    status_map = {"return": OrderStatus.returned, "exchange": OrderStatus.exchanged, "rejection": OrderStatus.rejected}
    if data.type in status_map:
        order.status = status_map[data.type]

    await db.commit()
    return {"message": "售后记录已创建"}

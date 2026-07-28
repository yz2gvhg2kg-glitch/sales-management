from typing import Optional
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
import io

from app.core.database import get_db
from app.core.security import get_current_user, get_current_admin
from app.models.customer import Customer, CustomerStatus
from app.models.user import User

router = APIRouter()


class CustomerCreate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    wechat: Optional[str] = None
    source: Optional[str] = None
    channel: Optional[str] = None
    source_product: Optional[str] = None
    remark: Optional[str] = None


class CustomerAssign(BaseModel):
    customer_ids: list[int]
    assigned_to: int


class CustomerStatusUpdate(BaseModel):
    status: CustomerStatus


@router.get("")
async def list_customers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    assigned_to: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Customer)
    count_query = select(func.count(Customer.id))

    # 普通员工只能看自己的客户
    if current_user.role == "employee":
        query = query.where(Customer.assigned_to == current_user.id)
        count_query = count_query.where(Customer.assigned_to == current_user.id)
    elif assigned_to:
        query = query.where(Customer.assigned_to == assigned_to)
        count_query = count_query.where(Customer.assigned_to == assigned_to)

    if status:
        query = query.where(Customer.status == status)
        count_query = count_query.where(Customer.status == status)
    if keyword:
        query = query.where(
            (Customer.name.contains(keyword)) | (Customer.phone.contains(keyword))
        )
        count_query = count_query.where(
            (Customer.name.contains(keyword)) | (Customer.phone.contains(keyword))
        )
    if start_date:
        query = query.where(Customer.created_at >= start_date)
        count_query = count_query.where(Customer.created_at >= start_date)
    if end_date:
        query = query.where(Customer.created_at <= end_date)
        count_query = count_query.where(Customer.created_at <= end_date)

    total = (await db.execute(count_query)).scalar()
    result = await db.execute(
        query.order_by(Customer.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )
    customers = result.scalars().all()

    return {
        "total": total,
        "items": [
            {
                "id": c.id,
                "name": c.name,
                "phone": c.phone,
                "wechat": c.wechat,
                "source": c.source,
                "channel": c.channel,
                "source_product": c.source_product,
                "assigned_to": c.assigned_to,
                "status": c.status.value,
                "assign_date": c.assign_date.isoformat() if c.assign_date else None,
                "add_date": c.add_date.isoformat() if c.add_date else None,
                "convert_date": c.convert_date.isoformat() if c.convert_date else None,
                "remark": c.remark,
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }
            for c in customers
        ],
    }


@router.post("")
async def create_customer(
    data: CustomerCreate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    customer = Customer(**data.model_dump())
    db.add(customer)
    await db.commit()
    await db.refresh(customer)
    return {"id": customer.id, "message": "创建成功"}


@router.post("/import")
async def import_customers(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    """批量导入客户数据（Excel）"""
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    customers = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        customer = Customer(
            name=row_data.get("姓名") or row_data.get("name"),
            phone=str(row_data.get("电话") or row_data.get("phone") or ""),
            wechat=row_data.get("微信") or row_data.get("wechat"),
            source=row_data.get("来源") or row_data.get("source"),
            channel=row_data.get("渠道") or row_data.get("channel"),
            source_product=row_data.get("引流产品") or row_data.get("source_product"),
            remark=row_data.get("备注") or row_data.get("remark"),
        )
        customers.append(customer)

    db.add_all(customers)
    await db.commit()
    return {"message": f"成功导入 {len(customers)} 条客户数据"}


@router.post("/assign")
async def assign_customers(
    data: CustomerAssign,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    """批量分配客户给业务员"""
    today = date.today()
    result = await db.execute(
        select(Customer).where(Customer.id.in_(data.customer_ids))
    )
    customers = result.scalars().all()

    for customer in customers:
        customer.assigned_to = data.assigned_to
        customer.status = CustomerStatus.assigned
        customer.assign_date = today

    await db.commit()
    return {"message": f"成功分配 {len(customers)} 个客户"}


@router.put("/{customer_id}/status")
async def update_customer_status(
    customer_id: int,
    data: CustomerStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新客户状态（业务员标记已添加等）"""
    result = await db.execute(select(Customer).where(Customer.id == customer_id))
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    # 普通员工只能操作自己的客户
    if current_user.role == "employee" and customer.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="无权操作此客户")

    today = date.today()
    customer.status = data.status
    if data.status == CustomerStatus.added:
        customer.add_date = today
    elif data.status == CustomerStatus.converted:
        customer.convert_date = today

    await db.commit()
    return {"message": "状态已更新"}

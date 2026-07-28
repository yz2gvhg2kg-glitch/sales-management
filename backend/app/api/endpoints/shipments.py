from typing import Optional
from datetime import date
from fastapi import APIRouter, Depends, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select, func, update
from sqlalchemy.ext.asyncio import AsyncSession
import io

from app.core.database import get_db
from app.core.security import get_current_admin
from app.models.order import Order, OrderStatus, Shipment

router = APIRouter()


class ShipmentCreate(BaseModel):
    order_ids: list[int]
    carrier: Optional[str] = None


@router.get("")
async def list_shipments(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    query = select(Shipment)
    count_query = select(func.count(Shipment.id))

    if keyword:
        query = query.where(Shipment.tracking_no.contains(keyword))
        count_query = count_query.where(Shipment.tracking_no.contains(keyword))

    total = (await db.execute(count_query)).scalar()
    result = await db.execute(
        query.order_by(Shipment.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )
    shipments = result.scalars().all()

    return {
        "total": total,
        "items": [
            {
                "id": s.id,
                "order_id": s.order_id,
                "tracking_no": s.tracking_no,
                "carrier": s.carrier,
                "ship_date": s.ship_date.isoformat() if s.ship_date else None,
                "status": s.status,
                "created_at": s.created_at.isoformat() if s.created_at else None,
            }
            for s in shipments
        ],
    }


@router.post("/batch-ship")
async def batch_ship(
    data: ShipmentCreate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    """批量发货"""
    today = date.today()
    shipments = []
    for order_id in data.order_ids:
        shipment = Shipment(
            order_id=order_id,
            carrier=data.carrier,
            ship_date=today,
        )
        shipments.append(shipment)

    db.add_all(shipments)

    # 更新订单状态为已发货
    await db.execute(
        update(Order).where(Order.id.in_(data.order_ids)).values(status=OrderStatus.shipped)
    )
    await db.commit()
    return {"message": f"成功发货 {len(data.order_ids)} 个订单"}


@router.post("/import-tracking")
async def import_tracking(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_admin),
):
    """批量导入快递单号"""
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        order_id = row_data.get("订单ID") or row_data.get("order_id")
        tracking_no = row_data.get("快递单号") or row_data.get("tracking_no")
        carrier = row_data.get("快递公司") or row_data.get("carrier")

        if order_id and tracking_no:
            result = await db.execute(
                select(Shipment).where(Shipment.order_id == int(order_id))
            )
            shipment = result.scalar_one_or_none()
            if shipment:
                shipment.tracking_no = str(tracking_no)
                if carrier:
                    shipment.carrier = str(carrier)
                count += 1

    await db.commit()
    return {"message": f"成功更新 {count} 条快递信息"}
